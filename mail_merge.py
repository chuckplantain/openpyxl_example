from openpyxl import load_workbook
import re


def segmentAdder(flightSegments, wb):
    '''
    :param flightSegments: a list of dicts containing flight info (only need length of list)
    :param wb: the workbook we are working on
    works with process merge. this function adds extra flight segment sections
    if necessary to a template xlsx file.
    TODO: programattically fill in the values
    for flight segments. Not sure if you had any thoughts on this.
    '''
    ws = wb.active
    pattern = re.compile(r'FLIGHT SEGMENT ')
    segments = (len(flightSegments))
    max_value = 55 + ( 23 * segments ) - 23
    for index in xrange(55, max_value, 23):
        crazy_inner_count = (index - 55) / 23 + 2
        pertinent_rows = ws['A32':'L54']
        blank_rows = ws.iter_rows(row_offset=index)
        for row, bl_row in zip(pertinent_rows, blank_rows):
            for cell, bl_cell in zip(row, bl_row):
                desired_style = cell.style
                bl_cell.style = desired_style
                if cell.value is not None:
                    val = cell.value
                    if pattern.search(val) is not None:
                        bl_cell.value = "Flight Segment " +  str(crazy_inner_count)
                    else:
                        bl_cell.value = cell.value
                else:
                    bl_cell.value = cell.value

    wb.save('/Users/kyleblazepetan/Desktop/foo.xlsx')


def process_merge(data, flightSegments):
    """
    :param data: a dict of values, keys being the <<Values>> in lowercase form
    :param flightSegmentData: array of dicts corresponding to flight legs
    :return: True if all went well
    """
    pattern = re.compile(r'(<<)(\w+)(>>)')
    wb = load_workbook('/Users/kyleblazepetan/Desktop/mailMergeTemplate.xlsx')
    ws = wb.active

    segmentAdder(flightSegments, wb)

    def lookUpFunction(matchobj):
        resp = data[str(matchobj.group(2)).lower()]
        if resp:
            return resp
        return 'key not found'

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.value = re.sub(pattern, lookUpFunction, cell.value.lower())

    wb.save('/Users/kyleblazepetan/Desktop/foo.xlsx')
