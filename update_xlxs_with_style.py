
# coding: utf-8

# In[1]:

from openpyxl import load_workbook
import re

data_dict = {}
data_dict['foo'] = 'Kyle Petan'
data_dict['bar'] = "Splendid and Great"

path_to_test_markedup_file = "/Users/kyleblazepetan/github/openpyxl_example/testFile.xlsx"
wb = load_workbook(path_to_test_markedup_file)
ws = wb.active

pattern = re.compile(r'(<<)(\w+)(>>)')

def lookUpFunction(matchobj):
    resp = data_dict[str(matchobj.group(2)).lower()]
    if resp:
        return resp
    return 'key not found'
    
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.value = re.sub(pattern, lookUpFunction, cell.value)
save_destination = "/Users/kyleblazepetan/github/openpyxl_example/new_shiny_file.xlsx"            
wb.save(save_destination)


# In[ ]:



