from openpyxl import load_workbook
import datetime

#
## Testing variables
template = '/Users/kyleblazepetan/Plumb/masterxlsx/INTERNATIONAL PASSENGER MANIFEST_MASTER.xlsx'
dict_list = [{u'date of birth': datetime.datetime(2015, 12, 13, 0, 0),
   u'expiration date': datetime.datetime(2015, 12, 13, 0, 0),
   u'first name': u'Kyle',
   u'issue date': datetime.datetime(2015, 12, 13, 0, 0),
   u'issuing country': u'United States of America',
   u'last name': u'Petan',
   u'middle initial': u'B',
   u'passport number': u'13213545454545554653',
   u'sex': u'M',
   u'weight': 150,
   u'weight unit (lbs/kg)': u'kg'},
  {u'date of birth': datetime.datetime(1930, 7, 17, 0, 0),
   u'expiration date': None,
   u'first name': u'Karen',
   u'issue date': None,
   u'issuing country': None,
   u'last name': u'Seleb',
   u'middle initial': u'Ann',
   u'passport number': None,
   u'sex': u'F',
   u'weight': 130,
   u'weight unit (lbs/kg)': u'lbs'},
  {u'date of birth': datetime.datetime(1999, 4, 12, 0, 0),
   u'expiration date': None,
   u'first name': u'Henry',
   u'issue date': None,
   u'issuing country': None,
   u'last name': u'Jones',
   u'middle initial': u'bill',
   u'passport number': None,
   u'sex': u'M',
   u'weight': 120,
   u'weight unit (lbs/kg)': u'kg'}]

trip_dict = {
u'date of trip': '12/24/2015',
u'trip number': 321456774,
u'generated on': '03/14/1998'
}

international = True

##
#

def write_xlsx_file(dict_list, trip_dict, template, international):
    '''
    :param dict_list: a list of dicts
    :param trip_dict: three item dict.
    :param template: uri for template xlsx file
    :param international: bool reflecting header style
    :return: True to indicate success
    '''
    wb = load_workbook(template)
    ws = wb.active
    header = []

    ws['C3'] = trip_dict[u'date of trip']
    ws['C4'] = trip_dict[u'trip number']
    ws['C8'] = trip_dict[u'generated on']

    if international:
        headerCells = ws['B11':'L11']
        cells = ws['B12':'L512']
    else:
        headerCells = ws['B11':'H11']
        cells = ws['B12':'H512']

    for row in headerCells:
        for cell in row:
            header.append(cell.value)

    for item, row in zip(dict_list, cells):
        for idx, cell in enumerate(row):
            if item[header[idx].lower()] is not None:
                cell.value = item[header[idx].lower()]
            else:
                continue

    wb.save('/Users/kyleblazepetan/Desktop/test.xlsx')
    return True
