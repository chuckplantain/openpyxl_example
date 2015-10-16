from openpyxl import load_workbook
import datetime

#
#CONSTANTS --- required fields
FIRST_NAME = 'FIRST NAME'
LAST_NAME = 'LAST NAME'
DATE_OF_BIRTH = 'DATE OF BIRTH'
SEX = 'SEX'
PASSPORT_NUMBER = 'PASSPORT NUMBER'
ISSUE_DATE = 'ISSUE DATE'
EXPIRATION_DATE = 'EXPIRATION DATE'
REQUIRED_FIELDS = [ FIRST_NAME, LAST_NAME, DATE_OF_BIRTH ]

def saveItem(item, state):
    for field in REQUIRED_FIELDS:
        try:
            val = item[field.lower()]
            if val is None:
                error_msg = 'ERROR: at least one required field is missing.  '
                error_msg += 'The Required Fields Are:'
                for fld in REQUIRED_FIELDS:
                    error_msg += "%s, " % fld
                return False, error_msg

        except:
            return False, 'The Header and dataFields are not the same'
    for key in item.keys():
        valid, data = verifyKeys(item, key)

        if (valid == False):
            return valid, data


    return True, item


def verifyKeys(item, key):
    if key == 'date of birth' or key == 'issue date' or key == 'expiration date':
        val = item[key]
        if val is not None:
            item[key] = datetime.datetime.strptime(val, "%m/%d/%Y")
            if not isinstance(item[key], datetime.date):
                return False, 'ERROR: dates should be in MM/DD/YYYY format please'



    if key == 'first name' or key == 'last name':
        val = item[key]
        if not (isinstance(val, unicode)):
            return False, 'ERROR: Names must be unicode and less than 55 characters long and are Required'

    if (key == 'weight'):
        val = item[key]
        if val is not None:
            if not (isinstance(val, int)):
                return False, 'ERROR: weight should be numerical'

    if key == 'weight unit':
        val = item[key]
        if val is not None:
            if not isinstance(val, unicode) or not (val.lower() == u'lb' or val.lower() == u'kg'):
                return False, 'ERROR: weight Unit field needs to be lb or kg'

    if key == 'sex':
        val = item[key]
        if val is not None:
            if not (val == 'm' or val == 'f' or val == 'M' or val == 'F'):
                return False, 'ERROR: gender should be specified with a m or a f'
    return True, 'No Errors'


def process_manifest(manifest_file):
    wb = load_workbook(manifest_file)
    ws = wb.active
    header = []
    data = []
    blankrow_count = 0

    for row in ws['B11':'L11']:
        for cell in row:
            if cell.value is not None:
                header.append(cell.value.lower())

    for row in ws['B12':'L512']:
        dataItem = {}
        for idx, cell in enumerate(row):
            try:
                key = header[idx]
            except:
                break
            else:
                dataItem[key] = cell.value
        if not all(v is None for v in dataItem.values()):
            valid, saved_item = saveItem(dataItem, data)
            if not valid:
                return False, saved_item
            else:
                data.append(saved_item)
        else:
            blankrow_count += 1
            if (blankrow_count >= 10):
                break
    return True, data
