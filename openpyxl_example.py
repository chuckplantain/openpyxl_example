
# coding: utf-8

# In[197]:

# import openpyxl as opxl
from openpyxl import load_workbook
import re

# test data dict
data_dict = {}
data_dict['foo'] = 'Kyle Petan'
data_dict['bar'] = "Splendid and Great"

# load a .xlxs file for practice
wb = load_workbook('./simpleTest2.xlsx')


# In[198]:

# regex to find strings matching <<*>>
# returns a 3-tuple, where we discard all but middle
pattern = re.compile(r'(<<)(\w+)(>>)')


# In[199]:

# work sheet
ws = wb.active

# Excel Files report correct rows and columns but libre office does not. 
# Something to lookinto
# ws.calculate_dimension()
# https://bitbucket.org/openpyxl/openpyxl/issues/278/get_highest_row-column-are-unreliable

# how many rows?
rows = ws.get_highest_row()
# how many columns?
cols = ws.get_highest_column()


# In[200]:


def lookUpFunction(matchobj):
    resp = data_dict[str(matchobj.group(2)).lower()]
    if resp:
        return resp
    return 'key_not_found!!!!'
    # OR ADD KEY IN Real life???

# In[201]:

for row in range(rows):
    for col in range(cols):
        currentCell = ws.cell(row = row, column = col)
        if currentCell.value is not None:
            currentCell.value = re.sub(pattern, lookUpFunction, currentCell.value)
            
wb.save('./sample.xlxs')
