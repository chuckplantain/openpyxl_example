
# coding: utf-8

# In[1]:

from openpyxl import load_workbook
import re

# test data dict
data_dict = {}
data_dict['foo'] = 'Kyle Petan'
data_dict['bar'] = "Splendid and Great"

# load a .xlxs file for practice
path_to_test_markedup_file = "Users/kyleblazepetan/github/openpyxl_example/testFile.xlsx"
wb = load_workbook(path_to_test_markedup_file)
ws = wb.active

# regex to find strings matching <<*>>
pattern = re.compile(r'(<<)(\w+)(>>)')

def lookUpFunction(matchobj):
    resp = data_dict[str(matchobj.group(2)).lower()]
    if resp:
        return resp
    return 'key_not_found!!!!'
    
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            # print(cell.value)
            cell.value = re.sub(pattern, lookUpFunction, cell.value)
edited_file_save_destination = "./new_shiny_file.xlsx"            
wb.save(edited_file_save_destination)


