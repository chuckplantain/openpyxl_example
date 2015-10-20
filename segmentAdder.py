from openpyxl import load_workbook
import re

wb = load_workbook('/Users/kyleblazepetan/Desktop/mailMergeTemplate.xlsx')
ws = wb.active

def segmentAdder(flightSegments, ws):
    pattern = re.compile(r'FLIGHT SEGMENT ')
    passedin = len(flightSegments)
    segments = 55 + ( 23 * passedin )
    for index in xrange(55, segments, 23):
        crazy_inner_count = (index - 55) / 23 + 2
        pertinent_rows = ws['A32':'L54']
        blank_rows = ws.iter_rows(row_offset=index)
        for row, bl_row in zip(pertinent_rows, blank_rows):
            for cell, bl_cell in zip(row, bl_row):
                desired_style = cell.style
                bl_cell.style = desired_style
                if cell.value is not None:
                    val = cell.value
                    if pattern.search(val) is not None:
                        bl_cell.value = "Flight Segment " +  str(crazy_inner_count)
                    else:
                        bl_cell.value = cell.value
                else:
                    bl_cell.value = cell.value

    wb.save('/Users/kyleblazepetan/Desktop/segements.xlsx')
