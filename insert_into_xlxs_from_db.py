from openpyxl import load_workbook
from manifest import process_manifest
import datetime

wb = load_workbook('/Users/kyleblazepetan/Plumb/xlsxFiles/INTERNATIONAL PASSENGER MANIFEST.xlsx')
ws = wb.active

input = '/Users/kyleblazepetan/Plumb/xlsxFiles/INTERNATIONAL PASSENGER MANIFEST.xlsx'
output = '/Users/kyleblazepetan/tmp/result.xlsx'
def write_xlsx_file(input, output):
    header = []
    data = process_manifest(input)
    for row in ws['D11':'N11']:
        for cell in row:
            header.append(cell.value)

    for item, row in zip(data, ws['D12':'N512']):
        for key, cell in zip(header, row):
            cell.value = item[key.lower()]

    wb.save(output)

write_xlsx_file(input, output)