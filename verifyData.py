
# coding: utf-8

# In[5]:

from openpyxl import load_workbook
import datetime


# In[6]:

header = []
data = []
path_to_file = "/Users/kyleblazepetan/Plumb/xlxsFiles/sampleFile.xlsx"
wb = load_workbook(path_to_file)
ws = wb.active

#CONSTANTS --- required fields
FIRST_NAME = 'FIRST NAME'
LAST_NAME = 'LAST NAME'
DATE_OF_BIRTH = 'DATE OF BIRTH'
SEX = 'SEX'
PASSPORT_NUMBER = 'PASSPORT NUMBER'
ISSUE_DATE = 'ISSUE DATE'
EXPIRATION_DATE = 'EXPIRATION DATE'
REQUIRED_FIELDS = [ FIRST_NAME, LAST_NAME, DATE_OF_BIRTH]


# In[7]:

def saveItem(item, state):
    for field in REQUIRED_FIELDS:        
        try:
            val = item[field.lower()]
        except:
            print('ERROR: at least one required field is missing')
            print 'Specifically: \n', item, '\nis missing ', field, ' field'
            print('The Required Fields Are:')
            for field in REQUIRED_FIELDS:
                print(field)              
    for key in item.keys():
        verifyKeys(item, key)
    state.append(item)

def verifyKeys(item, key):
    if (key == 'date of birth' or key == 'issue date' or key == 'expiration date'):
        val = item[key]
        if not isinstance(val, datetime.date):       
            print('ERROR: dates should be in MM/DD/YYYY format please')
        if (key == 'issue date' and item[key] > datetime.datetime.today()):
            print('ERROR: Passport has not been issued yet')
        if (key == 'expiration date' and item[key] < (datetime.datetime.today() or item['issue date'])):
            print('ERROR: Passport is expired')
            
    if (key == 'first name' or key == 'last name'):
        val = item[key]
        if not (isinstance(val, unicode)):
            print('ERROR: Names must be unicode and less than 55 characters long')
            
    if (key == 'weight'):
        val = item[key]
        if not (isinstance(val, int)):
            print('ERROR: weight should be numerical')
            
    if (key == 'sex'):
        val = item[key]
        if not (val == 'M' or 'F' or 'm' or 'f'):
            print('ERROR: gender should be specified with a M or a F')

            
for i, row in enumerate(ws['D11':'M511']):
    dataItem = {}
    for j, cell in enumerate(row):
        if (i <= 0):
            header.append(cell.value.lower())
        else:
            if cell.value is not None:
                dataItem[header[j]] = cell.value           
    if (dataItem):
        saveItem(dataItem, data)
print(data[1])


# In[ ]:



